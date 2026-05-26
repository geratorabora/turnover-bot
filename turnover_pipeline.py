# ===== 1A START =====
from __future__ import annotations  # 1A: современная типизация

import csv  # 1A: запись CSV без DBeaver
import sys  # 1A: подключение соседнего проекта csv_to_exel
from dataclasses import dataclass  # 1A: простой результат pipeline
from datetime import datetime  # 1A: дата отчёта для второго листа
from pathlib import Path  # 1A: пути к файлам
from typing import Optional  # 1A: опциональные значения

import pandas as pd  # 1A: batch-выгрузка в отдельный лист
import psycopg  # 1A: PostgreSQL
from openpyxl import Workbook  # 1A: отдельный xlsx только для вкладки по сериям
# ===== 1A END =====


# ===== 1B START =====
BOT_DIR = Path(__file__).resolve().parent  # 1B: папка turnover_bot
PROJECT_ROOT = BOT_DIR.parent  # 1B: общий корень проекта
LOCAL_EXPORT_SQL_PATH = BOT_DIR / "turnover_export.sql"  # 1B: SQL для Railway-репозитория turnover_bot
CSV_TO_EXEL_DIR = PROJECT_ROOT / "csv_to_exel"  # 1B: папка существующего Excel-конвертера
SHARED_EXPORT_SQL_PATH = CSV_TO_EXEL_DIR / "turnover_export.sql"  # 1B: SQL в соседнем локальном проекте
DEFAULT_EXPORT_SQL_PATH = LOCAL_EXPORT_SQL_PATH if LOCAL_EXPORT_SQL_PATH.exists() else SHARED_EXPORT_SQL_PATH  # 1B: выбираем доступный SQL
OUTPUT_XLSX_NAME = "turnover_pretty.xlsx"  # 1B: имя готового файла для пользователя
OUTPUT_BATCH_XLSX_NAME = "turnover_batch_stock.xlsx"  # 1B: отдельный файл по вкладке по сериям
OUTPUT_DISCREPANCIES_XLSX_NAME = "turnover_statement_discrepancies.xlsx"  # 1B: отдельный файл по расхождениям ведомости
OUTPUT_CSV_NAME = "turnover.csv"  # 1B: имя промежуточного CSV
# ===== 1B END =====


# ===== 1C START =====
if str(BOT_DIR) not in sys.path:  # 1C: сначала гарантируем приоритет локального модуля из turnover_bot
    sys.path.insert(0, str(BOT_DIR))  # 1C: добавляем текущую папку первой в sys.path
if CSV_TO_EXEL_DIR.exists() and str(CSV_TO_EXEL_DIR) not in sys.path:  # 1C: соседний проект оставляем только как запасной источник
    sys.path.append(str(CSV_TO_EXEL_DIR))  # 1C: добавляем его в конец, чтобы локальная копия имела приоритет

from csv_to_xlsx_turnover import (  # noqa: E402  # 1C: prettifier c приоритетом локальной версии
    add_batch_stock_analytics_sheet,
    add_batch_stock_sheet,
    add_statement_discrepancies_sheet,
    add_statement_discrepancies_summary_sheet,
    convert_turnover_csv_to_xlsx,
)
# ===== 1C END =====


# ===== 2A START =====
@dataclass
class TurnoverPipelineResult:
    csv_path: Path  # 2A: куда записали turnover.csv
    xlsx_path: Path  # 2A: куда записали turnover_pretty.xlsx
    batch_xlsx_path: Optional[Path]  # 2A: куда записали отдельный xlsx по сериям
    discrepancies_xlsx_path: Optional[Path]  # 2A: куда записали отдельный xlsx по расхождениям ведомости
    exported_rows: int  # 2A: сколько строк вернула SQL-выгрузка
# ===== 2A END =====


# ===== 2B START =====
def load_batch_stock_sheet_data(database_url: str, report_date: datetime) -> pd.DataFrame:
    """
    2B: Загружаем данные для дополнительного листа по сериям.
    Берём batch-остатки и подтягиваем среднюю себестоимость из turnover-таблицы.
    """

    sql = """
    with statement_qty as (
        select
            report_dt,
            item_code,
            sum(stock_qty) as statement_qty
        from public.raw_stock_statement
        where report_dt = %s
        group by report_dt, item_code
    ),
    cost_snapshot as (
        select
            report_dt,
            item_code,
            sum(stock_qty) as cost_qty,
            sum(stock_cost) as cost_total,
            case
                when nullif(sum(stock_qty), 0) is null then null
                else sum(stock_cost) / nullif(sum(stock_qty), 0)
            end as cost_unit
        from public.raw_stock_month_cost
        where report_dt = %s
        group by report_dt, item_code
    ),
    turnover_unit_cost as (
        select
            r.period::date as report_dt,
            r.item_code,
            max(trim(r.item)) as item,
            max(trim(r.article)) as article,
            coalesce(nullif(trim(max(r.article)), ''), r.item_code) as article_key,
            sum(coalesce(s.statement_qty, r.curr_stock_qty)) as turnover_qty,
            sum(
                case
                    when c.cost_qty is not null
                         and c.cost_unit is not null
                         and coalesce(s.statement_qty, r.curr_stock_qty) is not null
                        then case
                            when c.cost_qty >= coalesce(s.statement_qty, r.curr_stock_qty)
                                then c.cost_unit * coalesce(s.statement_qty, r.curr_stock_qty)
                            else c.cost_total + (
                                greatest(coalesce(s.statement_qty, r.curr_stock_qty) - c.cost_qty, 0)
                                * coalesce(r.curr_stock_cost / nullif(r.curr_stock_qty, 0), 0)
                            )
                        end
                    when s.statement_qty is not null and nullif(r.curr_stock_qty, 0) is not null
                        then (r.curr_stock_cost / nullif(r.curr_stock_qty, 0)) * s.statement_qty
                    else r.curr_stock_cost
                end
            ) as turnover_cost,
            case
                when nullif(sum(coalesce(s.statement_qty, r.curr_stock_qty)), 0) is null then null
                else sum(
                    case
                        when c.cost_qty is not null
                             and c.cost_unit is not null
                             and coalesce(s.statement_qty, r.curr_stock_qty) is not null
                            then case
                                when c.cost_qty >= coalesce(s.statement_qty, r.curr_stock_qty)
                                    then c.cost_unit * coalesce(s.statement_qty, r.curr_stock_qty)
                                else c.cost_total + (
                                    greatest(coalesce(s.statement_qty, r.curr_stock_qty) - c.cost_qty, 0)
                                    * coalesce(r.curr_stock_cost / nullif(r.curr_stock_qty, 0), 0)
                                )
                            end
                        when s.statement_qty is not null and nullif(r.curr_stock_qty, 0) is not null
                            then (r.curr_stock_cost / nullif(r.curr_stock_qty, 0)) * s.statement_qty
                        else r.curr_stock_cost
                    end
                ) / nullif(sum(coalesce(s.statement_qty, r.curr_stock_qty)), 0)
            end as avg_unit_cost
        from public.raw_turnover_stock r
        left join statement_qty s
            on s.report_dt = r.period::date
           and s.item_code = r.item_code
        left join cost_snapshot c
            on c.report_dt = r.period::date
           and c.item_code = r.item_code
        where r.period::date = %s
        group by r.period::date, r.item_code
    ),
    turnover_article_qty as (
        select
            article_key,
            sum(turnover_qty) as turnover_article_qty
        from turnover_unit_cost
        group by article_key
    ),
    batch_base as (
        select
            b.report_dt,
            b.item,
            b.item_code,
            b.article,
            coalesce(nullif(trim(b.article), ''), b.item_code) as article_key,
            b.quality,
            b.series,
            b.expiry_dt,
            b.batch_stock_qty,
            b.months_on_stock,
            b.estimated_prod_month
        from public.raw_stock_batches b
        where b.report_dt = %s
    ),
    batch_article_qty as (
        select
            article_key,
            sum(batch_stock_qty) as batch_article_qty
        from batch_base
        group by article_key
    )
    select
        coalesce(nullif(trim(b.item), ''), t.item) as "Наименование",
        coalesce(nullif(trim(b.article), ''), t.article) as "Артикул",
        b.quality as "Качество",
        b.series as "Серия",
        b.expiry_dt as "Годен до",
        b.batch_stock_qty as "Остаток по партиям",
        round(t.avg_unit_cost, 2) as "Средняя себестоимость",
        round(b.batch_stock_qty * t.avg_unit_cost, 2) as "Общая себестоимость",
        case
            when b.months_on_stock > 12 then 'свыше года'
            when b.months_on_stock >= 6 then 'от полгода до года'
            when b.months_on_stock >= 3 then 'от 3 месяцев до полгода'
            else 'менее 3 месяцев'
        end as "Уровень",
        b.months_on_stock as "Месяцев на складе",
        b.estimated_prod_month as "Оценочный месяц производства",
        ta.turnover_article_qty as turnover_article_qty,
        ba.batch_article_qty as batch_article_qty,
        ba.batch_article_qty - ta.turnover_article_qty as qty_diff
    from batch_base b
    left join turnover_unit_cost t
        on t.report_dt = b.report_dt
       and t.item_code = b.item_code
    left join turnover_article_qty ta
        on ta.article_key = b.article_key
    left join batch_article_qty ba
        on ba.article_key = b.article_key
    order by
        case
            when b.months_on_stock > 12 then 1
            when b.months_on_stock >= 6 then 2
            when b.months_on_stock >= 3 then 3
            else 4
        end,
        round(b.batch_stock_qty * t.avg_unit_cost, 2) desc nulls last,
        coalesce(nullif(trim(b.article), ''), t.article),
        b.quality,
        b.expiry_dt,
        b.series
    """

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set")

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (report_date.date(), report_date.date(), report_date.date(), report_date.date()))
            rows = cur.fetchall()
            columns = [col.name for col in cur.description]

    df = pd.DataFrame(rows, columns=columns)
    df = df.rename(
        columns={
            "turnover_article_qty": "Общее кол-во по артикулу в отчете по оборачиваемости",
            "batch_article_qty": "Общее кол-во по артикулу в отчете по сериям",
            "qty_diff": "Разница в количестве",
        }
    )
    return df
# ===== 2B END =====


# ===== 2C START =====
def export_batch_stock_xlsx(batch_stock_df: pd.DataFrame, xlsx_path: Path) -> Optional[Path]:
    """
    2C: Собираем отдельный xlsx только с листом по сериям.
    """

    if batch_stock_df is None or batch_stock_df.empty:
        return None

    wb = Workbook()
    wb.remove(wb.active)
    add_batch_stock_analytics_sheet(wb=wb, batch_stock_df=batch_stock_df)
    add_batch_stock_sheet(wb=wb, batch_stock_df=batch_stock_df)
    desired_order = ["Аналитика по сериям", "Остатки по сериям"]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]
    wb.save(xlsx_path)
    return xlsx_path
# ===== 2C END =====


# ===== 2CA START =====
def export_statement_discrepancies_xlsx(statement_discrepancies_df: pd.DataFrame, xlsx_path: Path) -> Optional[Path]:
    """
    2CA: Собираем отдельный xlsx со сверкой расхождений и сводкой по направлениям/сегментам.
    """

    if statement_discrepancies_df is None or statement_discrepancies_df.empty:
        return None

    wb = Workbook()
    wb.remove(wb.active)
    add_statement_discrepancies_sheet(wb=wb, statement_discrepancies_df=statement_discrepancies_df)
    add_statement_discrepancies_summary_sheet(wb=wb, statement_discrepancies_df=statement_discrepancies_df)
    desired_order = ["Перечень расхождений", "Сумма расхождений"]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]
    wb.save(xlsx_path)
    return xlsx_path
# ===== 2CA END =====


# ===== 2D START =====
def load_statement_adjustments(database_url: str, report_date: datetime) -> pd.DataFrame:
    """
    2D: Подготавливаем уточнение остатка по коду номенклатуры для листа детализации.
    """

    sql = """
    with statement_qty as (
        select
            report_dt,
            item_code,
            sum(stock_qty) as statement_qty
        from public.raw_stock_statement
        where report_dt = %s
        group by report_dt, item_code
    ),
    cost_snapshot as (
        select
            report_dt,
            item_code,
            sum(stock_qty) as cost_qty,
            sum(stock_cost) as cost_total,
            case
                when nullif(sum(stock_qty), 0) is null then null
                else sum(stock_cost) / nullif(sum(stock_qty), 0)
            end as cost_unit
        from public.raw_stock_month_cost
        where report_dt = %s
        group by report_dt, item_code
    )
    select
        r.item_code,
        coalesce(s.statement_qty, r.curr_stock_qty) as statement_qty,
        case
            when c.cost_qty is not null
                 and c.cost_unit is not null
                 and coalesce(s.statement_qty, r.curr_stock_qty) is not null
                then case
                    when c.cost_qty >= coalesce(s.statement_qty, r.curr_stock_qty)
                        then c.cost_unit * coalesce(s.statement_qty, r.curr_stock_qty)
                    else c.cost_total + (
                        greatest(coalesce(s.statement_qty, r.curr_stock_qty) - c.cost_qty, 0)
                        * coalesce(r.curr_stock_cost / nullif(r.curr_stock_qty, 0), 0)
                    )
                end
            when s.statement_qty is not null and nullif(r.curr_stock_qty, 0) is not null
                then (r.curr_stock_cost / nullif(r.curr_stock_qty, 0)) * s.statement_qty
            else r.curr_stock_cost
        end as adjusted_stock_cost
    from public.raw_turnover_stock r
    left join statement_qty s
        on s.report_dt = r.period::date
       and s.item_code = r.item_code
    left join cost_snapshot c
        on c.report_dt = r.period::date
       and c.item_code = r.item_code
    where r.period::date = %s
    """

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set")

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (report_date.date(), report_date.date(), report_date.date()))
            rows = cur.fetchall()
            columns = [col.name for col in cur.description]

    return pd.DataFrame(rows, columns=columns)
# ===== 2D END =====


# ===== 2DA START =====
def load_availability_adjustments(database_url: str, report_date: datetime) -> pd.DataFrame:
    """
    2DA: Подготавливаем "доступно сейчас" по коду номенклатуры для последней детализации.
    """

    sql = """
    select
        item_code,
        qty_available_now
    from public.raw_stock_availability
    where report_dt = %s
    """

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set")

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (report_date.date(),))
            rows = cur.fetchall()
            columns = [col.name for col in cur.description]

    return pd.DataFrame(rows, columns=columns)
# ===== 2DA END =====


# ===== 2E START =====
def load_statement_discrepancies(database_url: str, report_date: datetime) -> pd.DataFrame:
    """
    2E: Готовим перечень расхождений между остатком из отчёта по оборачиваемости
        и количеством из ведомости остатков.
    """

    sql = """
    with turnover_base as (
        select
            r.period::date as report_dt,
            max(trim(r.item)) as item,
            r.item_code,
            max(trim(r.article)) as article,
            max(trim(r.supplier)) as supplier,
            max(trim(r.segment)) as segment,
            max(trim(r.gau)) as direction,
            sum(r.curr_stock_qty) as turnover_qty,
            sum(r.curr_stock_cost) as turnover_cost,
            case
                when nullif(sum(r.curr_stock_qty), 0) is null then null
                else sum(r.curr_stock_cost) / nullif(sum(r.curr_stock_qty), 0)
            end as avg_unit_cost
        from public.raw_turnover_stock r
        where r.period::date = %s
        group by r.period::date, r.item_code
    ),
    statement_base as (
        select
            s.report_dt,
            s.item_code,
            sum(s.stock_qty) as statement_qty
        from public.raw_stock_statement s
        where s.report_dt = %s
        group by s.report_dt, s.item_code
    ),
    cost_snapshot as (
        select
            c.report_dt,
            c.item_code,
            sum(c.stock_qty) as cost_qty,
            sum(c.stock_cost) as cost_total,
            case
                when nullif(sum(c.stock_qty), 0) is null then null
                else sum(c.stock_cost) / nullif(sum(c.stock_qty), 0)
            end as cost_unit
        from public.raw_stock_month_cost c
        where c.report_dt = %s
        group by c.report_dt, c.item_code
    ),
    quality_base as (
        select
            b.report_dt,
            b.item_code,
            string_agg(distinct trim(b.quality), ', ' order by trim(b.quality)) as quality
        from public.raw_stock_batches b
        where b.report_dt = %s
        group by b.report_dt, b.item_code
    )
    select
        t.item as item_name,
        t.article as article,
        t.item_code as item_code,
        q.quality as quality,
        t.supplier as supplier,
        t.segment as segment,
        t.direction as direction,
        round(t.avg_unit_cost, 2) as turnover_avg_cost,
        t.turnover_qty as turnover_qty,
        round(t.turnover_cost, 2) as turnover_cost,
        s.statement_qty as statement_qty,
        round(
            case
                when s.statement_qty is null then null
                when c.cost_qty is not null and c.cost_unit is not null
                    then case
                        when c.cost_qty >= s.statement_qty
                            then c.cost_unit * s.statement_qty
                        else c.cost_total + (
                            greatest(s.statement_qty - c.cost_qty, 0)
                            * coalesce(t.avg_unit_cost, 0)
                        )
                    end
                when t.avg_unit_cost is not null
                    then s.statement_qty * t.avg_unit_cost
                else null
            end,
            2
        ) as statement_cost,
        coalesce(s.statement_qty, 0) - coalesce(t.turnover_qty, 0) as qty_diff,
        round(
            coalesce(
                case
                    when s.statement_qty is null then null
                    when c.cost_qty is not null and c.cost_unit is not null
                        then case
                            when c.cost_qty >= s.statement_qty
                                then c.cost_unit * s.statement_qty
                            else c.cost_total + (
                                greatest(s.statement_qty - c.cost_qty, 0)
                                * coalesce(t.avg_unit_cost, 0)
                            )
                        end
                    when t.avg_unit_cost is not null
                        then s.statement_qty * t.avg_unit_cost
                    else null
                end,
                0
            ) - coalesce(t.turnover_cost, 0),
            2
        ) as cost_diff
    from turnover_base t
    left join statement_base s
        on s.report_dt = t.report_dt
       and s.item_code = t.item_code
    left join cost_snapshot c
        on c.report_dt = t.report_dt
       and c.item_code = t.item_code
    left join quality_base q
        on q.report_dt = t.report_dt
       and q.item_code = t.item_code
    where s.statement_qty is distinct from t.turnover_qty
    order by
        abs(
            coalesce(
                case
                    when s.statement_qty is null then null
                    when c.cost_qty is not null and c.cost_unit is not null
                        then case
                            when c.cost_qty >= s.statement_qty
                                then c.cost_unit * s.statement_qty
                            else c.cost_total + (
                                greatest(s.statement_qty - c.cost_qty, 0)
                                * coalesce(t.avg_unit_cost, 0)
                            )
                        end
                    when t.avg_unit_cost is not null
                        then s.statement_qty * t.avg_unit_cost
                    else null
                end,
                0
            ) - coalesce(t.turnover_cost, 0)
        ) desc nulls last,
        abs(coalesce(s.statement_qty, 0) - coalesce(t.turnover_qty, 0)) desc nulls last,
        t.article,
        t.item
    """

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set")

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (report_date.date(), report_date.date(), report_date.date(), report_date.date()))
            rows = cur.fetchall()
            columns = [col.name for col in cur.description]

    df = pd.DataFrame(rows, columns=columns)
    return df.rename(
        columns={
            "item_name": "Номенклатура",
            "article": "Артикул",
            "item_code": "Код",
            "quality": "Качество",
            "supplier": "Поставщик",
            "segment": "Сегмент",
            "direction": "Направление",
            "turnover_avg_cost": "Средн себест из отчета по оборачиваемости",
            "turnover_qty": "Кол-во из отчета по оборачиваемости",
            "turnover_cost": "Себест из отчета по оборачиваемости",
            "statement_qty": "Кол-во из ведомости по остаткам",
            "statement_cost": "Сумма из ведомости по остаткам",
            "qty_diff": "Разница в кол-ве",
            "cost_diff": "Разница в себестоимости",
        }
    )
# ===== 2E END =====


# ===== 3A START =====
def export_turnover_csv(
    database_url: str,
    sql_path: Path,
    csv_path: Path,
) -> int:
    """
    3A: Выполняем SQL-выгрузку напрямую из Python и сохраняем long CSV.
    """

    if not database_url:  # 3A: без строки подключения работать нельзя
        raise RuntimeError("DATABASE_URL is not set")

    if not sql_path.exists():  # 3A: SQL должен быть рядом с проектом
        raise FileNotFoundError(f"SQL export file is missing: {sql_path}")

    sql = sql_path.read_text(encoding="utf-8")  # 3A: читаем SQL-выгрузку
    csv_path.parent.mkdir(parents=True, exist_ok=True)  # 3A: создаём временную папку, если нужно

    with psycopg.connect(database_url) as conn:  # 3A: подключаемся к PostgreSQL
        with conn.cursor() as cur:  # 3A: открываем курсор
            cur.execute(sql)  # 3A: выполняем SQL
            columns = [col.name for col in cur.description]  # 3A: берём имена колонок для CSV

            with csv_path.open("w", encoding="utf-8", newline="") as f:  # 3A: открываем CSV для записи
                writer = csv.writer(f, delimiter=";")  # 3A: пишем в формате, который ждёт конвертер
                writer.writerow(columns)  # 3A: первая строка = заголовки

                exported_rows = 0  # 3A: счётчик выгруженных строк
                for row in cur:  # 3A: потоково читаем результат SQL
                    writer.writerow(row)  # 3A: записываем строку в CSV
                    exported_rows += 1  # 3A: обновляем счётчик

    return exported_rows  # 3A: возвращаем размер выгрузки
# ===== 3A END =====


# ===== 4A START =====
def build_turnover_report(
    database_url: str,
    work_dir: Path,
    source_detail_path: Path,
    report_date: Optional[datetime] = None,
    include_batch_sheet: bool = False,
    sql_path: Optional[Path] = None,
) -> TurnoverPipelineResult:
    """
    4A: Полный pipeline после загрузки Excel в БД:
        DB SQL export -> turnover.csv -> turnover_pretty.xlsx.
    """

    export_sql_path = sql_path or DEFAULT_EXPORT_SQL_PATH  # 4A: используем SQL по умолчанию, если не передали другой
    csv_path = work_dir / OUTPUT_CSV_NAME  # 4A: промежуточный CSV во временной папке
    xlsx_path = work_dir / OUTPUT_XLSX_NAME  # 4A: итоговый Excel во временной папке
    batch_xlsx_path = work_dir / OUTPUT_BATCH_XLSX_NAME  # 4A: отдельный xlsx по серии
    discrepancies_xlsx_path = work_dir / OUTPUT_DISCREPANCIES_XLSX_NAME  # 4A: отдельный xlsx по расхождениям ведомости
    batch_stock_df = None  # 4A: по умолчанию дополнительного листа нет
    statement_adjustments_df = None  # 4A: уточнение количеств/себестоимости для детализации
    availability_adjustments_df = None  # 4A: уточнение свободного остатка для детализации
    statement_discrepancies_df = None  # 4A: отдельный лист со сверкой ведомости и оборачиваемости

    if report_date is not None:
        statement_adjustments_df = load_statement_adjustments(
            database_url=database_url,
            report_date=report_date,
        )
        availability_adjustments_df = load_availability_adjustments(
            database_url=database_url,
            report_date=report_date,
        )
        statement_discrepancies_df = load_statement_discrepancies(
            database_url=database_url,
            report_date=report_date,
        )
    if include_batch_sheet and report_date is not None:
        batch_stock_df = load_batch_stock_sheet_data(
            database_url=database_url,
            report_date=report_date,
        )

    exported_rows = export_turnover_csv(  # 4A: выполняем SQL и сохраняем CSV
        database_url=database_url,
        sql_path=export_sql_path,
        csv_path=csv_path,
    )

    convert_turnover_csv_to_xlsx(  # 4A: используем существующий Excel prettifier
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        source_detail_path=source_detail_path,
        batch_stock_df=batch_stock_df,
        statement_adjustments_df=statement_adjustments_df,
        availability_adjustments_df=availability_adjustments_df,
    )

    if not xlsx_path.exists():  # 4A: защита от тихого сбоя генерации
        raise RuntimeError(f"Final workbook was not created: {xlsx_path}")

    exported_batch_xlsx_path = None
    if include_batch_sheet and batch_stock_df is not None and not batch_stock_df.empty:
        exported_batch_xlsx_path = export_batch_stock_xlsx(
            batch_stock_df=batch_stock_df,
            xlsx_path=batch_xlsx_path,
        )

    exported_discrepancies_xlsx_path = export_statement_discrepancies_xlsx(
        statement_discrepancies_df=statement_discrepancies_df,
        xlsx_path=discrepancies_xlsx_path,
    )

    return TurnoverPipelineResult(  # 4A: возвращаем все важные пути и счётчик
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        batch_xlsx_path=exported_batch_xlsx_path,
        discrepancies_xlsx_path=exported_discrepancies_xlsx_path,
        exported_rows=exported_rows,
    )
# ===== 4A END =====
