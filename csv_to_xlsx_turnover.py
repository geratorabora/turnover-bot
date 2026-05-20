# ===== 1A START =====
from __future__ import annotations  # 1A: современная типизация

from pathlib import Path  # 1A: удобные пути к файлам
from typing import Dict, List, Optional, Tuple  # 1A: типы для читаемости
import re  # 1A: регулярки для распознавания колонок

import pandas as pd  # 1A: работа с CSV и таблицами

from openpyxl import load_workbook  # 1A: открываем xlsx после записи pandas
from openpyxl.styles import Alignment, Font, PatternFill  # 1A: стили ячеек, включая заливку
from openpyxl.utils import get_column_letter  # 1A: номер колонки -> буква Excel
from openpyxl.chart import LineChart, Reference  # 1A: графики Excel
# ===== 1A END =====


# ===== 1B START =====
INPUT_CSV_NAME = "turnover.csv"  # 1B: имя входного CSV по умолчанию
OUTPUT_XLSX_NAME = "turnover_pretty.xlsx"  # 1B: имя выходного Excel по умолчанию
SOURCE_DETAIL_XLSX_NAME = "96 для показателей КС (XLSX).xlsx"  # 1B: исходный файл 1С с детализацией по номенклатуре

SHEET_TOC = "Оглавление"  # 1B: первый лист с навигацией по файлу
SHEET_SUMMARY = "Итоги графики"  # 1B: лист с итоговыми графиками

SHEET_ALL = "общее"  # 1B: общий лист
SHEET_STOCK = "Общий остаток"  # 1B: лист по остаткам
SHEET_TURNS = "Оборачиваемость"  # 1B: лист по оборачиваемости
SHEET_SLOW = "Остаток НОТ"  # 1B: лист по низкооборачиваемым товарам
SHEET_NONLIQ = "Остаток неликвидов"  # 1B: лист по неликвидам
SHEET_BATCH_STOCK = "Остатки по сериям"  # 1B: лист со вторым отчётом по сериям и качеству
SHEET_DETAIL_LAST_WEEK = "детализация последняя неделя"  # 1B: последний лист с исходной детализацией, имя <= 31 символа

SHEET_CHART_STOCK = "График общий остаток"  # 1B: график остатков
SHEET_CHART_TURNS = "График Оборачиваемость"  # 1B: график оборачиваемости
SHEET_CHART_SLOW = "График остаток НОТ"  # 1B: график низкооборачиваемых
SHEET_CHART_NONLIQ = "График остаток неликвидов"  # 1B: график неликвидов
# ===== 1B END =====


# ===== 2A START =====
def normalize_numeric(series: pd.Series) -> pd.Series:
    """
    2A: Аккуратно переводим серию в numeric.
        Всё нечисловое превращаем в NaN, потом в None.
    """
    converted = pd.to_numeric(series, errors="coerce")  # 2A: пробуем перевести в число
    converted = converted.where(pd.notnull(converted), None)  # 2A: NaN меняем на None
    return converted  # 2A: возвращаем очищенную серию
# ===== 2A END =====


# ===== 2B START =====
def load_long_csv(csv_path: Path, sep: str, encoding: str) -> pd.DataFrame:
    """
    2B: Читаем длинный CSV и приводим типы к порядку.

    ВАЖНО:
    - нормализуем текстовые ключи pg / segment / pg_segment
    - убираем неразрывные пробелы
    - убираем хвостовые пробелы
    - схлопываем двойные пробелы внутри текста
    - для сегментов сохраняем ведущие 3 пробела в pg_segment,
      потому что они нужны нам для визуального отличия сегментов от PG
    """

    df = pd.read_csv(csv_path, sep=sep, encoding=encoding)  # 2B: читаем CSV

    required_cols = {
        "pg",
        "segment",
        "lvl",
        "pg_segment",
        "week_dt",
        "week_num",
        "stock_cost",
        "sales_cost",
        "av_stock_cost",
        "turns_rub",
        "slow_stock_lt2",
        "nonliq_stock",
    }  # 2B: обязательные колонки нового длинного CSV

    missing = sorted(list(required_cols - set(df.columns)))  # 2B: проверяем, все ли нужные колонки есть
    if missing:  # 2B: если чего-то не хватает
        raise ValueError(f"В CSV не найдены обязательные колонки: {missing}")  # 2B: бросаем понятную ошибку

    df = df.copy()  # 2B: делаем копию, чтобы безопасно менять значения

    def normalize_text_value(value) -> str | None:
        """
        2B: Нормализуем обычный текстовый ключ:
            - None/NaN -> None
            - неразрывные пробелы -> обычные
            - trim по краям
            - двойные пробелы внутри -> один пробел
        """
        if pd.isna(value):  # 2B: пустые значения оставляем пустыми
            return None

        text = str(value)  # 2B: приводим к строке
        text = text.replace("\u00A0", " ").replace("\u202F", " ")  # 2B: заменяем неразрывные пробелы
        text = text.strip()  # 2B: убираем пробелы по краям
        text = " ".join(text.split())  # 2B: схлопываем повторяющиеся пробелы внутри
        return text if text != "" else None  # 2B: пустую строку превращаем в None

    def normalize_pg_segment_value(value, lvl_value) -> str | None:
        """
        2B: Нормализуем pg_segment.
            Для PG:
                "Онкогенетика" -> "Онкогенетика"
            Для сегмента:
                "   Микробиология ..." -> "   Микробиология ..."
            То есть ведущие 3 пробела для сегментов сохраняем специально.
        """
        cleaned = normalize_text_value(value)  # 2B: сначала чистим текст как обычный ключ
        if cleaned is None:  # 2B: если после очистки пусто
            return None

        try:
            lvl_num = int(float(lvl_value)) if not pd.isna(lvl_value) else None  # 2B: аккуратно приводим lvl к числу
        except Exception:
            lvl_num = None  # 2B: если не удалось, считаем уровень неизвестным

        if lvl_num == 2:  # 2B: если это сегмент
            return "   " + cleaned  # 2B: возвращаем сегмент с ведущими 3 пробелами
        return cleaned  # 2B: если это PG, просто возвращаем очищенный текст

    df["week_dt"] = pd.to_datetime(df["week_dt"], errors="coerce")  # 2B: дату недели приводим к datetime
    df["week_num"] = pd.to_numeric(df["week_num"], errors="coerce")  # 2B: номер недели приводим к числу
    df["lvl"] = pd.to_numeric(df["lvl"], errors="coerce")  # 2B: уровень тоже приводим к числу

    df["pg"] = df["pg"].apply(normalize_text_value)  # 2B: чистим ключ PG
    df["segment"] = df["segment"].apply(normalize_text_value)  # 2B: чистим ключ segment
    df["pg_segment"] = [
        normalize_pg_segment_value(pg_segment_value, lvl_value)
        for pg_segment_value, lvl_value in zip(df["pg_segment"], df["lvl"])
    ]  # 2B: чистим ключ pg_segment с учётом уровня строки

    df["stock_cost"] = normalize_numeric(df["stock_cost"])  # 2B: остаток -> число
    df["sales_cost"] = normalize_numeric(df["sales_cost"])  # 2B: расход -> число
    df["av_stock_cost"] = normalize_numeric(df["av_stock_cost"])  # 2B: средний остаток -> число
    df["turns_rub"] = normalize_numeric(df["turns_rub"])  # 2B: оборачиваемость -> число
    df["slow_stock_lt2"] = normalize_numeric(df["slow_stock_lt2"])  # 2B: низкооборачиваемый остаток -> число
    df["nonliq_stock"] = normalize_numeric(df["nonliq_stock"])  # 2B: неликвидный остаток -> число

    df = df.sort_values(
        by=["week_num", "pg", "lvl", "pg_segment"],
        ascending=[True, True, True, True],
        na_position="last",
    ).reset_index(drop=True)  # 2B: базовая сортировка для предсказуемости

    return df  # 2B: возвращаем подготовленный длинный датафрейм
# ===== 2B END =====


# ===== 3A START =====
def build_week_map(df: pd.DataFrame) -> Dict[int, str]:
    """
    3A: Строим карту:
        week_num -> метка недели для колонок Excel
    Пример:
        1 -> w01
        2 -> w02
        11 -> w11
    """

    week_nums = sorted([int(x) for x in df["week_num"].dropna().unique()])  # 3A: получаем уникальные номера недель

    week_map: Dict[int, str] = {}  # 3A: сюда запишем отображение

    for week_num in week_nums:  # 3A: идём по всем неделям
        week_map[week_num] = f"w{week_num:02d}"  # 3A: создаём метку недели

    return week_map  # 3A: возвращаем словарь
# ===== 3A END =====


# ===== 3B START =====
def build_sort_order(df: pd.DataFrame) -> List[str]:
    """
    3B: Строим правильный порядок строк:
        - сначала PG
        - потом его сегменты
        - якорь сортировки = stock_cost на последней неделе
    """

    max_week_num = int(df["week_num"].max())  # 3B: определяем последнюю неделю в данных

    last_week_df = df[df["week_num"] == max_week_num].copy()  # 3B: берём только последнюю неделю

    pg_order_df = (
        last_week_df[last_week_df["lvl"] == 1][["pg_segment", "stock_cost"]]
        .sort_values(by=["stock_cost", "pg_segment"], ascending=[False, True], na_position="last")
        .reset_index(drop=True)
    )  # 3B: сортируем PG по остатку на последней неделе

    final_order: List[str] = []  # 3B: сюда будем собирать итоговый порядок строк

    for _, pg_row in pg_order_df.iterrows():  # 3B: идём по всем PG в нужном порядке
        pg_label = pg_row["pg_segment"]  # 3B: название PG
        final_order.append(pg_label)  # 3B: сначала добавляем сам PG

        seg_order_df = (
            last_week_df[(last_week_df["lvl"] == 2) & (last_week_df["pg"] == pg_label)][["pg_segment", "stock_cost"]]
            .sort_values(by=["stock_cost", "pg_segment"], ascending=[False, True], na_position="last")
            .reset_index(drop=True)
        )  # 3B: сегменты данного PG сортируем тоже по остатку на последней неделе

        for _, seg_row in seg_order_df.iterrows():  # 3B: добавляем сегменты после PG
            final_order.append(seg_row["pg_segment"])  # 3B: добавляем сегмент в итоговый список

    return final_order  # 3B: возвращаем порядок строк
# ===== 3B END =====


# ===== 4A START =====
def build_wide_metric_table(
    df: pd.DataFrame,
    value_col: str,
    prefix_label: str,
    row_order: List[str],
    week_map: Dict[int, str],
) -> pd.DataFrame:
    """
    4A: Строим wide-таблицу для одной метрики.
    На выходе:
        pg_segment | <метрика w01> | <метрика w02> | ...
    """

    pivot_df = df.pivot_table(
        index="pg_segment",
        columns="week_num",
        values=value_col,
        aggfunc="first",
    )  # 4A: разворачиваем длинную таблицу в широкую по неделям

    existing_week_nums = sorted([int(x) for x in pivot_df.columns.tolist()])  # 4A: получаем реально существующие недели

    ordered_week_cols = existing_week_nums  # 4A: недели уже упорядочены по возрастанию

    pivot_df = pivot_df.reindex(row_order)  # 4A: переставляем строки в нужном порядке
    pivot_df = pivot_df.reindex(columns=ordered_week_cols)  # 4A: переставляем колонки по порядку недель

    new_column_names = {}  # 4A: сюда запишем новые имена колонок

    for week_num in ordered_week_cols:  # 4A: идём по всем неделям
        new_column_names[week_num] = f"{prefix_label} {week_map[week_num]}"  # 4A: формируем подпись колонки

    pivot_df = pivot_df.rename(columns=new_column_names)  # 4A: переименовываем числовые week_num в человекочитаемые названия
    pivot_df = pivot_df.reset_index()  # 4A: возвращаем pg_segment из индекса обратно в колонку

    return pivot_df  # 4A: возвращаем готовую wide-таблицу
# ===== 4A END =====


# ===== 4B START =====
def build_all_sheet(
    stock_df: pd.DataFrame,
    turns_df: pd.DataFrame,
    slow_df: pd.DataFrame,
    nonliq_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    4B: Собираем общий лист:
        pg_segment + чередование
        Остаток w01, Об. руб w01, Низкооб w01, Неликвиды w01, ...
    """

    result_df = stock_df[["pg_segment"]].copy()  # 4B: начинаем с первой колонки

    stock_cols = [c for c in stock_df.columns if c != "pg_segment"]  # 4B: колонки остатка
    turns_cols = [c for c in turns_df.columns if c != "pg_segment"]  # 4B: колонки оборачиваемости
    slow_cols = [c for c in slow_df.columns if c != "pg_segment"]  # 4B: колонки низкооборачиваемых
    nonliq_cols = [c for c in nonliq_df.columns if c != "pg_segment"]  # 4B: колонки неликвидов

    week_labels = []  # 4B: сюда соберём список недельных хвостов вроде w01, w02

    for col in stock_cols:  # 4B: смотрим на колонки остатка
        match = re.match(r"^Остаток (w\d+)$", col)  # 4B: вытаскиваем часть wNN
        if match:  # 4B: если совпало
            week_labels.append(match.group(1))  # 4B: сохраняем метку недели

    for week_label in week_labels:  # 4B: собираем общий лист по одной неделе за раз
        stock_col = f"Остаток {week_label}"  # 4B: имя колонки остатка
        turns_col = f"Об. руб {week_label}"  # 4B: имя колонки оборачиваемости
        slow_col = f"Низкооб {week_label}"  # 4B: имя колонки низкооборачиваемых
        nonliq_col = f"Неликвиды {week_label}"  # 4B: имя колонки неликвидов

        if stock_col in stock_df.columns:  # 4B: если колонка есть
            result_df[stock_col] = stock_df[stock_col]  # 4B: добавляем её
        if turns_col in turns_df.columns:  # 4B: если колонка есть
            result_df[turns_col] = turns_df[turns_col]  # 4B: добавляем её
        if slow_col in slow_df.columns:  # 4B: если колонка есть
            result_df[slow_col] = slow_df[slow_col]  # 4B: добавляем её
        if nonliq_col in nonliq_cols:  # 4B: если колонка есть
            result_df[nonliq_col] = nonliq_df[nonliq_col]  # 4B: добавляем её

    return result_df  # 4B: возвращаем общий лист
# ===== 4B END =====

# ===== 4C START =====
def build_summary_metrics_table(df_long: pd.DataFrame, week_map: Dict[int, str]) -> pd.DataFrame:
    """
    4C: Строим маленькую таблицу для общего листа графиков.

    На выходе получаем по одной строке на неделю:
    - week_label
    - total_stock
    - total_slow
    - total_nonliq
    - total_turns

    ВАЖНО:
    total_turns считаем правильно:
        sum(sales_cost) / sum(av_stock_cost)
    причём только по строкам PG (lvl = 1),
    чтобы не задвоить агрегаты через сегменты.
    """

    pg_only = df_long[df_long["lvl"] == 1].copy()  # 4C: берём только строки PG, чтобы не было двойного счёта

    summary_df = (
        pg_only.groupby(["week_num", "week_dt"], as_index=False)
        .agg(
            total_stock=("stock_cost", "sum"),
            total_slow=("slow_stock_lt2", "sum"),
            total_nonliq=("nonliq_stock", "sum"),
            total_sales=("sales_cost", "sum"),
            total_av_stock=("av_stock_cost", "sum"),
        )
        .sort_values(by=["week_num", "week_dt"], ascending=[True, True])
        .reset_index(drop=True)
    )  # 4C: агрегируем по неделе все нужные суммы

    summary_df["total_turns"] = summary_df.apply(
        lambda row: (row["total_sales"] / row["total_av_stock"]) if pd.notna(row["total_av_stock"]) and row["total_av_stock"] not in (0, 0.0) else None,
        axis=1,
    )  # 4C: считаем общую оборачиваемость недели

    summary_df["week_label"] = summary_df["week_num"].apply(lambda x: week_map[int(x)])  # 4C: делаем подпись недели вида w01

    summary_df = summary_df[
        ["week_num", "week_dt", "week_label", "total_stock", "total_slow", "total_nonliq", "total_turns"]
    ].copy()  # 4C: оставляем только нужные колонки

    return summary_df  # 4C: возвращаем итоговую маленькую таблицу
# ===== 4C END =====


# ===== 5A START =====
def write_data_sheets(
    xlsx_path: Path,
    full_df: pd.DataFrame,
    stock_df: pd.DataFrame,
    turns_df: pd.DataFrame,
    slow_df: pd.DataFrame,
    nonliq_df: pd.DataFrame,
) -> None:
    """
    5A: Записываем табличные листы в Excel.
    """

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:  # 5A: открываем writer
        full_df.to_excel(writer, index=False, sheet_name=SHEET_ALL)  # 5A: общий лист
        stock_df.to_excel(writer, index=False, sheet_name=SHEET_STOCK)  # 5A: лист остатков
        turns_df.to_excel(writer, index=False, sheet_name=SHEET_TURNS)  # 5A: лист оборачиваемости
        slow_df.to_excel(writer, index=False, sheet_name=SHEET_SLOW)  # 5A: лист низкооборачиваемых
        nonliq_df.to_excel(writer, index=False, sheet_name=SHEET_NONLIQ)  # 5A: лист неликвидов
# ===== 5A END =====


# ===== 5B START =====
def format_worksheet(ws, first_col_width: float = 52) -> None:
    """
    5B: Универсально форматируем табличный лист.

    Логика:
    - строки PG (верхний уровень) делаем жирными и светло-серыми
    - строки сегментов оставляем обычными, с отступом
    - числовые форматы задаём по типу колонок
    """

    header_font = Font(bold=True)  # 5B: жирный шрифт заголовка
    pg_font = Font(bold=True)  # 5B: жирный шрифт для строк PG
    base_align = Alignment(vertical="center", wrap_text=False)  # 5B: обычное выравнивание
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)  # 5B: заголовки переносим по словам
    seg_align = Alignment(vertical="center", indent=2)  # 5B: выравнивание с отступом для сегментов
    pg_align = Alignment(vertical="center")  # 5B: выравнивание для PG
    pg_fill = PatternFill(fill_type="solid", fgColor="EDEDED")  # 5B: светло-серая заливка для строк PG

    ws.freeze_panes = "B2"  # 5B: закрепляем верхнюю строку и первую колонку
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"  # 5B: включаем фильтр

    headers = [cell.value for cell in ws[1]]  # 5B: читаем заголовки
    col_idx = {h: i + 1 for i, h in enumerate(headers)}  # 5B: делаем словарь заголовков

    for cell in ws[1]:  # 5B: форматируем строку заголовка
        cell.font = header_font  # 5B: делаем жирной
        cell.alignment = header_align  # 5B: включаем переносы по словам

    fmt_int_thousands = "#,##0"  # 5B: формат для целых значений
    fmt_turns = "0.00"  # 5B: формат для оборачиваемости

    for header in headers:  # 5B: идём по всем заголовкам
        if not isinstance(header, str):  # 5B: нестроковые пропускаем
            continue

        col_letter = get_column_letter(col_idx[header])  # 5B: вычисляем букву колонки

        if re.match(r"^Остаток w\d+$", header):  # 5B: если колонка остатка
            ws.column_dimensions[col_letter].width = 14  # 5B: ставим ширину
            for row_num in range(2, ws.max_row + 1):  # 5B: форматируем все ячейки данных
                ws.cell(row_num, col_idx[header]).number_format = fmt_int_thousands  # 5B: формат целого

        elif re.match(r"^Низкооб w\d+$", header):  # 5B: если колонка низкооборачиваемых
            ws.column_dimensions[col_letter].width = 14  # 5B: ставим ширину
            for row_num in range(2, ws.max_row + 1):  # 5B: форматируем значения
                ws.cell(row_num, col_idx[header]).number_format = fmt_int_thousands  # 5B: формат целого

        elif re.match(r"^Неликвиды w\d+$", header):  # 5B: если колонка неликвидов
            ws.column_dimensions[col_letter].width = 14  # 5B: ставим ширину
            for row_num in range(2, ws.max_row + 1):  # 5B: форматируем значения
                ws.cell(row_num, col_idx[header]).number_format = fmt_int_thousands  # 5B: формат целого

        elif re.match(r"^Об\. руб w\d+$", header):  # 5B: если колонка оборачиваемости
            ws.column_dimensions[col_letter].width = 12  # 5B: ставим ширину
            for row_num in range(2, ws.max_row + 1):  # 5B: форматируем значения
                ws.cell(row_num, col_idx[header]).number_format = fmt_turns  # 5B: формат дробного

    ws.row_dimensions[1].height = 32  # 5B: даём шапке место под переносы

    ws.column_dimensions["A"].width = first_col_width  # 5B: делаем первую колонку широкой

    for row_num in range(2, ws.max_row + 1):  # 5B: идём по строкам данных
        first_cell = ws.cell(row_num, 1)  # 5B: первая ячейка строки
        value = first_cell.value  # 5B: её значение

        is_segment = isinstance(value, str) and value.startswith("   ")  # 5B: сегменты у нас начинаются с пробелов

        if is_segment:  # 5B: если это сегмент
            first_cell.alignment = seg_align  # 5B: ставим отступ
        else:  # 5B: если это строка PG
            first_cell.font = pg_font  # 5B: делаем жирной
            first_cell.alignment = pg_align  # 5B: обычное выравнивание для PG

            for col_num in range(1, ws.max_column + 1):  # 5B: красим всю строку PG
                ws.cell(row_num, col_num).fill = pg_fill  # 5B: светло-серая заливка

        for col_num in range(2, ws.max_column + 1):  # 5B: выравнивание для числовых ячеек
            ws.cell(row_num, col_num).alignment = base_align  # 5B: обычное выравнивание
# ===== 5B END =====


# ===== 6A START =====
def add_line_chart_sheet(
    wb,
    source_sheet_name: str,
    chart_sheet_name: str,
    chart_title: str,
    anchor_cell: str = "A2",
) -> None:
    """
    6A: Строим лист с графиком по табличному листу.

    Логика:
    - на график берём только строки верхнего уровня (PG)
    - сегменты с ведущими пробелами пропускаем
    - подписи рядов берём из первой колонки
    - легенду ставим снизу
    - для рублёвых графиков:
        * подписи оси Y в "М руб"
        * шаг оси Y = 25 млн
        * ось X проходит через 0
    - для графика оборачиваемости:
        * обычный числовой масштаб
        * ось X тоже проходит через 0
    """

    source_ws = wb[source_sheet_name]  # 6A: лист-источник с табличными данными
    chart_ws = wb.create_sheet(chart_sheet_name)  # 6A: создаём новый лист под график

    chart = LineChart()  # 6A: создаём линейный график
    chart.title = chart_title  # 6A: задаём заголовок
    chart.style = 2  # 6A: стандартный стиль Excel
    chart.height = 28  # 6A: делаем график высоким
    chart.width = 30  # 6A: нормальная ширина
    chart.legend.position = "b"  # 6A: легенду ставим вниз

    chart.x_axis.delete = False  # 6A: ось X оставляем видимой
    chart.y_axis.delete = False  # 6A: ось Y оставляем видимой

    categories = Reference(
        source_ws,
        min_col=2,
        min_row=1,
        max_col=source_ws.max_column,
        max_row=1,
    )  # 6A: подписи по оси X = заголовки недель

    included_values = []  # 6A: сюда соберём все числовые значения рядов, попавших на график

    for row_num in range(2, source_ws.max_row + 1):  # 6A: идём по строкам листа
        row_label = source_ws.cell(row=row_num, column=1).value  # 6A: читаем подпись строки из 1-й колонки

        if isinstance(row_label, str) and row_label.startswith("   "):
            continue  # 6A: сегменты пропускаем, на графике оставляем только PG

        row_ref = Reference(
            source_ws,
            min_col=1,
            min_row=row_num,
            max_col=source_ws.max_column,
            max_row=row_num,
        )  # 6A: строка целиком: [название ряда | значения]

        chart.add_data(
            row_ref,
            titles_from_data=True,  # 6A: первая ячейка строки = имя ряда
            from_rows=True,         # 6A: каждая строка = отдельный ряд
        )

        for col_num in range(2, source_ws.max_column + 1):  # 6A: собираем числовые значения ряда
            cell_value = source_ws.cell(row=row_num, column=col_num).value  # 6A: берём значение ячейки
            if isinstance(cell_value, (int, float)):  # 6A: если это число
                included_values.append(float(cell_value))  # 6A: добавляем его в список

    chart.set_categories(categories)  # 6A: назначаем подписи оси X

    if included_values:  # 6A: если есть данные для расчёта масштаба
        max_val = max(included_values)  # 6A: максимальное значение на графике

        # 6A: отдельная настройка для графика оборачиваемости
        if source_sheet_name == SHEET_TURNS:
            upper_bound = max_val * 1.10 if max_val > 0 else 1  # 6A: небольшой запас сверху
            chart.y_axis.scaling.min = -0.25 * upper_bound  # 6A: создаём нижнюю зону под легенду
            chart.y_axis.scaling.max = upper_bound  # 6A: верхняя граница
            chart.y_axis.numFmt = '0.00'  # 6A: обычный числовой формат для оборачиваемости
            chart.x_axis.crosses = "autoZero"  # 6A: ось X проходит через 0

        # 6A: настройка для рублёвых графиков
        else:
            major_unit = 25_000_000  # 6A: шаг оси Y = 25 млн руб

            if max_val <= 0:  # 6A: защита на случай пустых/нулевых значений
                upper_bound = major_unit  # 6A: минимальная верхняя граница
            else:
                upper_bound = int((max_val + major_unit - 1) // major_unit) * major_unit  # 6A: округляем вверх до шага 25 млн

            chart.y_axis.scaling.min = -major_unit  # 6A: даём одну "ступеньку" ниже нуля под легенду
            chart.y_axis.scaling.max = upper_bound  # 6A: верхняя граница оси
            chart.y_axis.majorUnit = major_unit  # 6A: шаг основной сетки = 25 млн
            chart.y_axis.numFmt = '0,," М руб"'  # 6A: подписи оси Y показываем в миллионах рублей
            chart.x_axis.crosses = "autoZero"  # 6A: ось X проходит через 0 и визуально выделяет нулевую линию

    chart_ws.add_chart(chart, anchor_cell)  # 6A: вставляем график на лист
    chart_ws.sheet_view.showGridLines = False  # 6A: убираем сетку листа
# ===== 6A END =====

# ===== 6B START =====
def add_summary_dashboard_sheet(wb, summary_df: pd.DataFrame) -> None:
    """
    6B: Создаём лист с четырьмя общими графиками:
    1. динамика общего остатка
    2. динамика общего остатка низкооборачиваемых товаров
    3. динамика общего остатка неликвидов
    4. динамика общей оборачиваемости

    На листе также оставляем небольшую служебную таблицу-источник данных.

    Масштаб оси Y:
    - нижняя граница = 70% от минимального положительного значения ряда
    - верхняя граница = округление вверх до удобного шага
    """

    ws = wb.create_sheet(SHEET_SUMMARY)  # 6B: создаём новый лист

    # 6B: Пишем служебную таблицу, от которой будут строиться графики
    ws["A1"] = "Неделя"
    ws["B1"] = "Общий остаток"
    ws["C1"] = "Общий остаток низкооборачиваемых товаров (НОТ)"
    ws["D1"] = "Общий остаток неликвидов"
    ws["E1"] = "Общая оборачиваемость"

    for idx, row in summary_df.iterrows():  # 6B: переносим данные по неделям в лист
        excel_row = idx + 2  # 6B: данные начинаются со 2-й строки
        ws.cell(excel_row, 1).value = row["week_label"]  # 6B: подпись недели
        ws.cell(excel_row, 2).value = row["total_stock"]  # 6B: общий остаток
        ws.cell(excel_row, 3).value = row["total_slow"]  # 6B: общий НОТ
        ws.cell(excel_row, 4).value = row["total_nonliq"]  # 6B: общий неликвид
        ws.cell(excel_row, 5).value = row["total_turns"]  # 6B: общая оборачиваемость

    # 6B: Форматируем служебную таблицу
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for row_num in range(2, ws.max_row + 1):
        ws.cell(row_num, 2).number_format = '#,##0'
        ws.cell(row_num, 3).number_format = '#,##0'
        ws.cell(row_num, 4).number_format = '#,##0'
        ws.cell(row_num, 5).number_format = '0.00'

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.row_dimensions[1].height = 40

    categories = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_col=1,
        max_row=ws.max_row,
    )  # 6B: категории по оси X = недели

    def calc_axis_bounds(series: pd.Series, major_unit: float, min_factor: float = 0.70) -> tuple[float, float]:
        """
        6B: Считаем нижнюю и верхнюю границы оси Y.

        Правило:
        - lower = 70% от минимального положительного значения
        - upper = округление вверх до шага major_unit
        """
        values = pd.to_numeric(series, errors="coerce").dropna()  # 6B: оставляем только числа

        if values.empty:  # 6B: если данных нет
            return 0.0, float(major_unit)  # 6B: безопасный диапазон

        positive_values = values[values > 0]  # 6B: берём только положительные значения

        if positive_values.empty:  # 6B: если положительных значений нет
            return 0.0, float(major_unit)  # 6B: безопасный диапазон

        min_val = float(positive_values.min())  # 6B: минимальное положительное значение
        max_val = float(positive_values.max())  # 6B: максимальное положительное значение

        lower_bound = min_val * min_factor  # 6B: нижняя граница = 70% от минимума
        upper_bound = ((max_val + major_unit - 1) // major_unit) * major_unit if major_unit >= 1 else None  # 6B: шаблон для крупных шагов

        if major_unit >= 1:  # 6B: для рублёвых графиков шаг большой, округляем как целое
            upper_bound = float(int((max_val + major_unit - 1) // major_unit) * major_unit)
        else:  # 6B: для дробного шага 0.25 делаем обычное округление вверх
            upper_bound = float((int(max_val / major_unit) + 1) * major_unit)

        if upper_bound <= lower_bound:  # 6B: защита от слишком узкого диапазона
            upper_bound = lower_bound + major_unit  # 6B: добавляем один шаг сверху

        return float(lower_bound), float(upper_bound)  # 6B: возвращаем границы

    major_unit_rub = 25_000_000  # 6B: шаг по оси Y для рублёвых графиков = 25 млн
    major_unit_turns = 0.25  # 6B: шаг по оси Y для оборачиваемости = 0.25

    stock_lower, stock_upper = calc_axis_bounds(summary_df["total_stock"], major_unit_rub)  # 6B: границы для общего остатка
    slow_lower, slow_upper = calc_axis_bounds(summary_df["total_slow"], major_unit_rub)  # 6B: границы для общего НОТ
    nonliq_lower, nonliq_upper = calc_axis_bounds(summary_df["total_nonliq"], major_unit_rub)  # 6B: границы для общего неликвида
    turns_lower, turns_upper = calc_axis_bounds(summary_df["total_turns"], major_unit_turns)  # 6B: границы для общей оборачиваемости

    # ---------- график 1: общий остаток ----------
    chart_stock = LineChart()
    chart_stock.title = "Динамика общего остатка"
    chart_stock.style = 2
    chart_stock.height = 10
    chart_stock.width = 24
    chart_stock.legend = None

    data_stock = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_col=2,
        max_row=ws.max_row,
    )
    chart_stock.add_data(data_stock, titles_from_data=True)
    chart_stock.set_categories(categories)

    chart_stock.x_axis.delete = False  # 6B: ось X показываем
    chart_stock.y_axis.delete = False  # 6B: ось Y показываем
    chart_stock.x_axis.title = "Недели"  # 6B: подпись оси X
    chart_stock.y_axis.title = "М руб"  # 6B: подпись оси Y
    chart_stock.y_axis.majorUnit = major_unit_rub  # 6B: шаг по оси Y = 25 млн
    chart_stock.y_axis.scaling.min = stock_lower  # 6B: нижняя граница = 70% от минимума
    chart_stock.y_axis.scaling.max = stock_upper  # 6B: верхняя граница оси
    chart_stock.y_axis.numFmt = '0,," М руб"'  # 6B: подписи в миллионах рублей

    ws.add_chart(chart_stock, "G2")

    # ---------- график 2: общий НОТ ----------
    chart_slow = LineChart()
    chart_slow.title = "Динамика остатка низкооборачиваемых товаров (НОТ)"
    chart_slow.style = 2
    chart_slow.height = 10
    chart_slow.width = 24
    chart_slow.legend = None

    data_slow = Reference(
        ws,
        min_col=3,
        min_row=1,
        max_col=3,
        max_row=ws.max_row,
    )
    chart_slow.add_data(data_slow, titles_from_data=True)
    chart_slow.set_categories(categories)

    chart_slow.x_axis.delete = False  # 6B: ось X показываем
    chart_slow.y_axis.delete = False  # 6B: ось Y показываем
    chart_slow.x_axis.title = "Недели"  # 6B: подпись оси X
    chart_slow.y_axis.title = "М руб"  # 6B: подпись оси Y
    chart_slow.y_axis.majorUnit = major_unit_rub  # 6B: шаг по оси Y = 25 млн
    chart_slow.y_axis.scaling.min = slow_lower  # 6B: нижняя граница = 70% от минимума
    chart_slow.y_axis.scaling.max = slow_upper  # 6B: верхняя граница оси
    chart_slow.y_axis.numFmt = '0,," М руб"'  # 6B: подписи в миллионах рублей

    ws.add_chart(chart_slow, "G22")

    # ---------- график 3: общий неликвид ----------
    chart_nonliq = LineChart()
    chart_nonliq.title = "Динамика остатка неликвидов"
    chart_nonliq.style = 2
    chart_nonliq.height = 10
    chart_nonliq.width = 24
    chart_nonliq.legend = None

    data_nonliq = Reference(
        ws,
        min_col=4,
        min_row=1,
        max_col=4,
        max_row=ws.max_row,
    )
    chart_nonliq.add_data(data_nonliq, titles_from_data=True)
    chart_nonliq.set_categories(categories)

    chart_nonliq.x_axis.delete = False  # 6B: ось X показываем
    chart_nonliq.y_axis.delete = False  # 6B: ось Y показываем
    chart_nonliq.x_axis.title = "Недели"  # 6B: подпись оси X
    chart_nonliq.y_axis.title = "М руб"  # 6B: подпись оси Y
    chart_nonliq.y_axis.majorUnit = major_unit_rub  # 6B: шаг по оси Y = 25 млн
    chart_nonliq.y_axis.scaling.min = nonliq_lower  # 6B: нижняя граница = 70% от минимума
    chart_nonliq.y_axis.scaling.max = nonliq_upper  # 6B: верхняя граница оси
    chart_nonliq.y_axis.numFmt = '0,," М руб"'  # 6B: подписи в миллионах рублей

    ws.add_chart(chart_nonliq, "G42")

    # ---------- график 4: общая оборачиваемость ----------
    chart_turns = LineChart()
    chart_turns.title = "Динамика общей оборачиваемости"
    chart_turns.style = 2
    chart_turns.height = 10
    chart_turns.width = 24
    chart_turns.legend = None

    data_turns = Reference(
        ws,
        min_col=5,
        min_row=1,
        max_col=5,
        max_row=ws.max_row,
    )
    chart_turns.add_data(data_turns, titles_from_data=True)
    chart_turns.set_categories(categories)

    chart_turns.x_axis.delete = False  # 6B: ось X показываем
    chart_turns.y_axis.delete = False  # 6B: ось Y показываем
    chart_turns.x_axis.title = "Недели"  # 6B: подпись оси X
    chart_turns.y_axis.title = "Об."  # 6B: подпись оси Y
    chart_turns.y_axis.majorUnit = major_unit_turns  # 6B: шаг по оси Y = 0.25
    chart_turns.y_axis.scaling.min = turns_lower  # 6B: нижняя граница = 70% от минимума
    chart_turns.y_axis.scaling.max = turns_upper  # 6B: верхняя граница оси
    chart_turns.y_axis.numFmt = '0.00'  # 6B: формат чисел по оси Y

    ws.add_chart(chart_turns, "G62")
# ===== 6B END =====

# ===== 6C START =====
def add_toc_sheet(wb) -> None:
    """
    6C: Создаём лист-оглавление.

    На листе:
    - список всех рабочих листов файла
    - кликабельная ссылка на каждый лист
    - краткое описание, что на нём находится
    """

    ws = wb.create_sheet(SHEET_TOC)  # 6C: создаём новый лист "Оглавление"

    # 6C: Заголовок листа
    ws["A1"] = "Оглавление отчёта"
    ws["A1"].font = Font(bold=True, size=14)

    # 6C: Шапка таблицы
    ws["A3"] = "Лист"
    ws["B3"] = "Описание"
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(vertical="center", wrap_text=True)

    # 6C: Описания для всех листов отчёта
    descriptions = {
        SHEET_SUMMARY: "Четыре итоговых графика: динамика общего остатка, НОТ, неликвидов и общей оборачиваемости по неделям.",
        SHEET_ALL: "Сводная таблица по всем продукт-группам и сегментам: остаток, оборачиваемость, НОТ и неликвиды по всем неделям.",
        SHEET_STOCK: "Таблица только по общему остатку в разрезе продукт-групп и сегментов.",
        SHEET_TURNS: "Таблица только по оборачиваемости в разрезе продукт-групп и сегментов.",
        SHEET_SLOW: "Таблица только по низкооборачиваемым товарам (НОТ).",
        SHEET_NONLIQ: "Таблица только по остаткам неликвидных товаров.",
        SHEET_BATCH_STOCK: "Детализация по сериям и качеству: остаток по партиям, средняя и общая себестоимость, месяцы на складе.",
        SHEET_CHART_STOCK: "График динамики общего остатка по продукт-группам.",
        SHEET_CHART_TURNS: "График динамики оборачиваемости по продукт-группам.",
        SHEET_CHART_SLOW: "График динамики низкооборачиваемых товаров по продукт-группам.",
        SHEET_CHART_NONLIQ: "График динамики остатков неликвидных товаров по продукт-группам.",
        SHEET_DETAIL_LAST_WEEK: "Детализация до номенклатуры по последней неделе. Если нужно посмотреть остаток и свободный в себестоимости в разрезе сегмента, поставщика, направления или менеджера, это можно сделать тут.",
    }

    # 6C: Порядок листов, который хотим показать в оглавлении
    ordered_sheet_names = [
        SHEET_SUMMARY,
        SHEET_ALL,
        SHEET_STOCK,
        SHEET_TURNS,
        SHEET_SLOW,
        SHEET_NONLIQ,
        SHEET_BATCH_STOCK,
        SHEET_CHART_STOCK,
        SHEET_CHART_TURNS,
        SHEET_CHART_SLOW,
        SHEET_CHART_NONLIQ,
        SHEET_DETAIL_LAST_WEEK,
    ]

    current_row = 4  # 6C: данные начинаются с 4-й строки

    for sheet_name in ordered_sheet_names:  # 6C: идём по листам в нужном порядке
        if sheet_name not in wb.sheetnames:  # 6C: если листа нет, пропускаем
            continue

        cell = ws.cell(current_row, 1)  # 6C: ячейка под имя листа
        cell.value = sheet_name  # 6C: пишем имя листа
        cell.hyperlink = f"#'{sheet_name}'!A1"  # 6C: делаем внутреннюю ссылку на лист
        cell.style = "Hyperlink"  # 6C: применяем стиль гиперссылки

        desc_cell = ws.cell(current_row, 2)  # 6C: ячейка под описание листа
        desc_cell.value = descriptions.get(sheet_name, "")  # 6C: пишем описание листа
        desc_cell.alignment = Alignment(vertical="top", wrap_text=True)  # 6C: включаем перенос по словам
        if sheet_name == SHEET_DETAIL_LAST_WEEK:  # 6C: описание детализации делаем жирным
            desc_cell.font = Font(bold=True)  # 6C: выделяем прикладное описание

        current_row += 1  # 6C: переходим на следующую строку

    # 6C: Немного форматируем ширины колонок
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 4

    # 6C: Закрепляем верх таблицы
    ws.freeze_panes = "A4"
# ===== 6C END =====


# ===== 6D START =====
def add_last_week_detail_sheet(wb, source_detail_path: Path) -> None:
    """
    6D: Добавляем в итоговый файл последний лист с детализацией до номенклатуры.

    Логика:
    - берём внешний файл 1С из папки проекта
    - читаем его первый лист
    - переносим все данные как есть
    - удаляем самую нижнюю строку, если это строка итогов
    - добавляем лист в итоговый файл последним
    """

    if not source_detail_path.exists():  # 6D: если файла нет, тихо выходим
        return

    source_wb = load_workbook(source_detail_path, data_only=False)  # 6D: открываем исходный файл 1С
    source_ws = source_wb[source_wb.sheetnames[0]]  # 6D: берём первый лист исходного файла

    if SHEET_DETAIL_LAST_WEEK in wb.sheetnames:  # 6D: если лист уже есть, удаляем старый вариант
        del wb[SHEET_DETAIL_LAST_WEEK]

    target_ws = wb.create_sheet(SHEET_DETAIL_LAST_WEEK)  # 6D: создаём новый лист в итоговом файле
    columns_to_drop = {  # 6D: набор колонок, которые на листе детализации не нужны
        "Номенклатура.Код",
        "Средний остаток, шт",
        "Расход, шт",
        "Выручка",
        "Себестоимость продаж за период",
        "Себестоимость среднего остатка",
        "Вал.Пр",
        "Рент. %",
        "Рент.Тов.Зап",
    }
    money_columns = {  # 6D: колонки, где хотим включить отображение с разделением разрядов
        "Конечный остаток (товары)",
        "Себестоимость (из отч. себ)",
        "Свободный остаток текущий",
        "Себестоимость свободного остатка",
        "Рзв",
        "Себ.Рзв",
    }

    # 6D: Копируем значения ячеек построчно и поколоночно
    for row in source_ws.iter_rows():  # 6D: идём по всем строкам исходного листа
        for cell in row:  # 6D: идём по всем ячейкам строки
            target_ws[cell.coordinate].value = cell.value  # 6D: переносим значение ячейки

    # 6D: Определяем, похожа ли последняя строка на строку итогов
    if target_ws.max_row >= 2:  # 6D: есть смысл проверять только если строк больше одной
        last_row = target_ws.max_row  # 6D: номер последней строки

        row_values = [
            target_ws.cell(last_row, col_num).value
            for col_num in range(1, target_ws.max_column + 1)
        ]  # 6D: собираем значения последней строки

        row_values_as_text = " ".join(
            str(v).strip().lower()
            for v in row_values
            if v is not None and str(v).strip() != ""
        )  # 6D: склеиваем текст последней строки для простой проверки

        # 6D: Если в последней строке есть слово "итог", "итоги", "итого" или "всего",
        #      считаем её итоговой и удаляем
        if any(marker in row_values_as_text for marker in ["итог", "итоги", "итого", "всего"]):
            target_ws.delete_rows(last_row, 1)  # 6D: удаляем последнюю строку

    header_values = [target_ws.cell(1, col_num).value for col_num in range(1, target_ws.max_column + 1)]  # 6D: читаем заголовки
    cols_to_delete = [  # 6D: вычисляем номера колонок, которые нужно убрать
        idx
        for idx, header_value in enumerate(header_values, start=1)
        if isinstance(header_value, str) and header_value.strip() in columns_to_drop
    ]

    for col_num in reversed(cols_to_delete):  # 6D: удаляем справа налево, чтобы индексы не сдвигались
        target_ws.delete_cols(col_num, 1)  # 6D: физически удаляем ненужную колонку

    header_font = Font(bold=True)  # 6D: шрифт шапки
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)  # 6D: переносы в шапке
    data_align = Alignment(vertical="top", wrap_text=False)  # 6D: обычное выравнивание данных
    fmt_int_thousands = "#,##0"  # 6D: формат с разделением разрядов для стоимостных и количественных колонок
    odd_row_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # 6D: белая заливка для нечётных строк данных
    even_row_fill = PatternFill(fill_type="solid", fgColor="E7E7E7")  # 6D: более заметная светло-серая заливка для чётных строк данных

    for col_num in range(1, target_ws.max_column + 1):  # 6D: идём по всем колонкам после удаления
        col_letter = target_ws.cell(1, col_num).column_letter  # 6D: получаем букву колонки
        target_ws.column_dimensions[col_letter].width = 10  # 6D: ставим одинаковую максимальную ширину

        if col_num == 1:  # 6D: колонку "Номенклатура" делаем шире остальных
            target_ws.column_dimensions[col_letter].width = 30  # 6D: даём больше места под название товара

        header_cell = target_ws.cell(1, col_num)  # 6D: ячейка заголовка
        header_cell.font = header_font  # 6D: делаем заголовок жирным
        header_cell.alignment = header_align  # 6D: включаем переносы по словам
        header_value = header_cell.value.strip() if isinstance(header_cell.value, str) else header_cell.value  # 6D: нормализуем имя заголовка

        if header_value == "Period":  # 6D: колонку периода скрываем, но не удаляем
            target_ws.column_dimensions[col_letter].hidden = True  # 6D: скрываем колонку периода

        for row_num in range(2, target_ws.max_row + 1):  # 6D: форматируем строки данных
            data_cell = target_ws.cell(row_num, col_num)  # 6D: текущая ячейка данных
            data_cell.alignment = data_align  # 6D: оставляем данные без переноса
            data_cell.fill = odd_row_fill if row_num % 2 == 0 else even_row_fill  # 6D: чередуем белый и бледно-серый фон
            if header_value in money_columns:  # 6D: для нужных колонок включаем разделение разрядов
                data_cell.number_format = fmt_int_thousands  # 6D: применяем числовой формат

    target_ws.auto_filter.ref = f"A1:{target_ws.cell(1, target_ws.max_column).column_letter}{target_ws.max_row}"  # 6D: включаем фильтр на весь диапазон
    target_ws.row_dimensions[1].height = 36  # 6D: даём шапке место под переносы
    target_ws.freeze_panes = "A2"  # 6D: закрепляем верхнюю строку
    target_ws.sheet_view.showGridLines = False  # 6D: убираем стандартную сетку, чтобы заливка читалась чище
# ===== 6D END =====


# ===== 6E START =====
def add_batch_stock_sheet(wb, batch_stock_df: pd.DataFrame) -> None:
    """
    6E: Добавляем отдельный лист по остаткам в разрезе серий/качеств.
    """

    if batch_stock_df is None or batch_stock_df.empty:
        return

    if SHEET_BATCH_STOCK in wb.sheetnames:
        del wb[SHEET_BATCH_STOCK]

    ws = wb.create_sheet(SHEET_BATCH_STOCK)

    for col_num, column_name in enumerate(batch_stock_df.columns, start=1):
        cell = ws.cell(1, col_num)
        cell.value = column_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for row_num, row in enumerate(batch_stock_df.itertuples(index=False), start=2):
        fill = PatternFill(fill_type="solid", fgColor="FFFFFF" if row_num % 2 == 0 else "E7E7E7")
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row_num, col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"
    ws.row_dimensions[1].height = 32

    width_map = {
        "A": 36,
        "B": 14,
        "C": 18,
        "D": 26,
        "E": 14,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 20,
        "K": 18,
        "L": 18,
        "M": 16,
    }

    money_columns = {"Средняя себестоимость", "Общая себестоимость"}
    qty_columns = {
        "Остаток по партиям",
        "Общее кол-во по артикулу в отчете по оборачиваемости",
        "Общее кол-во по артикулу в отчете по сериям",
        "Разница в количестве",
    }

    for col_num in range(1, ws.max_column + 1):
        col_letter = ws.cell(1, col_num).column_letter
        ws.column_dimensions[col_letter].width = width_map.get(col_letter, 14)
        header_value = ws.cell(1, col_num).value

        for data_row in range(2, ws.max_row + 1):
            cell = ws.cell(data_row, col_num)
            if header_value in money_columns:
                cell.number_format = "#,##0.00"
            elif header_value in qty_columns:
                cell.number_format = "#,##0"
            elif header_value == "Годен до":
                cell.number_format = "dd.mm.yyyy"
            elif header_value == "Оценочный месяц производства":
                cell.number_format = "mm.yyyy"
# ===== 6E END =====


# ===== 7A START =====
def convert_turnover_csv_to_xlsx(
    csv_path: str | Path,
    xlsx_path: str | Path | None = None,
    sep: str = ";",
    encoding: str = "utf-8",
    source_detail_path: Optional[str | Path] = None,
    batch_stock_df: Optional[pd.DataFrame] = None,
) -> Path:
    """
    7A: Главная функция:
        читаем длинный CSV ->
        строим wide-таблицы ->
        пишем Excel ->
        форматируем ->
        строим графики ->
        добавляем итоговый лист и оглавление ->
        добавляем детализацию по номенклатуре за последнюю неделю ->
        переставляем листы в нужный порядок
    """

    csv_path = Path(csv_path)  # 7A: приводим входной путь к Path

    if xlsx_path is None:  # 7A: если путь к xlsx не передан
        xlsx_path = csv_path.with_suffix(".xlsx")  # 7A: создаём рядом с csv
    else:  # 7A: если путь передан явно
        xlsx_path = Path(xlsx_path)  # 7A: тоже приводим его к Path

    if source_detail_path is None:  # 7A: если файл детализации не передан явно
        detail_path = Path(SOURCE_DETAIL_XLSX_NAME)  # 7A: сохраняем прежнее поведение локального скрипта
    else:  # 7A: если файл детализации передан извне
        detail_path = Path(source_detail_path)  # 7A: используем загруженный пользователем Excel

    df_long = load_long_csv(csv_path, sep=sep, encoding=encoding)  # 7A: читаем длинный CSV

    week_map = build_week_map(df_long)  # 7A: строим карту недель
    row_order = build_sort_order(df_long)  # 7A: строим порядок строк
    summary_df = build_summary_metrics_table(df_long, week_map)  # 7A: строим таблицу для общего листа графиков

    stock_df = build_wide_metric_table(
        df=df_long,
        value_col="stock_cost",
        prefix_label="Остаток",
        row_order=row_order,
        week_map=week_map,
    )  # 7A: wide-таблица по остаткам

    turns_df = build_wide_metric_table(
        df=df_long,
        value_col="turns_rub",
        prefix_label="Об. руб",
        row_order=row_order,
        week_map=week_map,
    )  # 7A: wide-таблица по оборачиваемости

    slow_df = build_wide_metric_table(
        df=df_long,
        value_col="slow_stock_lt2",
        prefix_label="Низкооб",
        row_order=row_order,
        week_map=week_map,
    )  # 7A: wide-таблица по низкооборачиваемым остаткам

    nonliq_df = build_wide_metric_table(
        df=df_long,
        value_col="nonliq_stock",
        prefix_label="Неликвиды",
        row_order=row_order,
        week_map=week_map,
    )  # 7A: wide-таблица по неликвидам

    full_df = build_all_sheet(
        stock_df=stock_df,
        turns_df=turns_df,
        slow_df=slow_df,
        nonliq_df=nonliq_df,
    )  # 7A: собираем общий лист

    write_data_sheets(
        xlsx_path=xlsx_path,
        full_df=full_df,
        stock_df=stock_df,
        turns_df=turns_df,
        slow_df=slow_df,
        nonliq_df=nonliq_df,
    )  # 7A: записываем табличные листы

    wb = load_workbook(xlsx_path)  # 7A: открываем Excel для форматирования и графиков

    format_worksheet(wb[SHEET_ALL], first_col_width=52)  # 7A: форматируем общий лист
    format_worksheet(wb[SHEET_STOCK], first_col_width=52)  # 7A: форматируем остатки
    format_worksheet(wb[SHEET_TURNS], first_col_width=52)  # 7A: форматируем оборачиваемость
    format_worksheet(wb[SHEET_SLOW], first_col_width=52)  # 7A: форматируем низкооборачиваемые
    format_worksheet(wb[SHEET_NONLIQ], first_col_width=52)  # 7A: форматируем неликвиды

    add_line_chart_sheet(
        wb=wb,
        source_sheet_name=SHEET_STOCK,
        chart_sheet_name=SHEET_CHART_STOCK,
        chart_title="Общий остаток на складах отгрузки",
    )  # 7A: график остатков

    add_line_chart_sheet(
        wb=wb,
        source_sheet_name=SHEET_TURNS,
        chart_sheet_name=SHEET_CHART_TURNS,
        chart_title="Оборачиваемость",
    )  # 7A: график оборачиваемости

    add_line_chart_sheet(
        wb=wb,
        source_sheet_name=SHEET_SLOW,
        chart_sheet_name=SHEET_CHART_SLOW,
        chart_title="Остатки низкооборачиваемых товаров (НОТ)",
    )  # 7A: график низкооборачиваемых

    add_line_chart_sheet(
        wb=wb,
        source_sheet_name=SHEET_NONLIQ,
        chart_sheet_name=SHEET_CHART_NONLIQ,
        chart_title="Остатки неликвидов",
    )  # 7A: график неликвидов

    add_summary_dashboard_sheet(
        wb=wb,
        summary_df=summary_df,
    )  # 7A: добавляем новый лист с итоговыми графиками

    add_last_week_detail_sheet(
        wb=wb,
        source_detail_path=detail_path,
    )  # 7A: добавляем лист с детализацией до номенклатуры за последнюю неделю

    add_batch_stock_sheet(
        wb=wb,
        batch_stock_df=batch_stock_df,
    )  # 7A: при наличии добавляем лист по сериям и качеству

    add_toc_sheet(wb=wb)  # 7A: добавляем лист-оглавление

    # 7A: Переставляем листы в нужный итоговый порядок
    desired_order = [
        SHEET_TOC,
        SHEET_SUMMARY,
        SHEET_ALL,
        SHEET_STOCK,
        SHEET_TURNS,
        SHEET_SLOW,
        SHEET_NONLIQ,
        SHEET_BATCH_STOCK,
        SHEET_CHART_STOCK,
        SHEET_CHART_TURNS,
        SHEET_CHART_SLOW,
        SHEET_CHART_NONLIQ,
        SHEET_DETAIL_LAST_WEEK,
    ]  # 7A: желаемый порядок листов

    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]  # 7A: физически переставляем листы

    wb.save(xlsx_path)  # 7A: сохраняем итоговый файл

    return xlsx_path  # 7A: возвращаем путь к готовому Excel
# ===== 7A END =====


# ===== 8A START =====
if __name__ == "__main__":  # 8A: запуск локального скрипта
    output_file = convert_turnover_csv_to_xlsx(
        INPUT_CSV_NAME,  # 8A: входной CSV
        OUTPUT_XLSX_NAME,  # 8A: выходной Excel
    )
    print(f"Saved: {output_file}")  # 8A: печатаем путь к готовому файлу
# ===== 8A END =====
